import openpyxl
import os
from openpyxl.drawing.image import Image

def add_rows_around_color_codes(excel_path, new_excel_path, image_folder):
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    # Identify rows with color codes in column A starting from row 11
    color_code_rows = []
    for row in sheet.iter_rows(min_row=11, min_col=1, max_col=1, values_only=False):
        cell = row[0]
        if cell.value:
            color_code_rows.append(cell.row)

    # Function to find an image in all subfolders by matching the first 4 digits of the filename
    def find_image(color_code, base_folder):
        color_code_prefix = str(color_code)[:4]
        for root, dirs, files in os.walk(base_folder):
            for file in files:
                if file.startswith(color_code_prefix):
                    return os.path.join(root, file)
        return None

    # Add rows around each color code or group of color codes
    offset = 0
    i = 0
    while i < len(color_code_rows):
        current_row = color_code_rows[i]

        # Determine the end of the current group of consecutive rows
        group_end = current_row
        while i + 1 < len(color_code_rows) and color_code_rows[i + 1] == group_end + 1:
            group_end += 1
            i += 1

        # Adjust the positions with the offset
        start_with_offset = current_row + offset
        end_with_offset = group_end + offset

        # Insert two rows above the group
        sheet.insert_rows(start_with_offset, amount=2)
        offset += 2
        start_with_offset += 2
        end_with_offset += 2

        # Insert two rows below the group
        sheet.insert_rows(end_with_offset + 1, amount=2)
        offset += 2

        # Merge cells in column E for the current group
        merge_start_row = start_with_offset
        merge_end_row = min(start_with_offset + 4, end_with_offset + 4)
        sheet.merge_cells(f'E{merge_start_row}:E{merge_end_row}')

        # Calculate the total height of the merged rows
        total_height = 0
        for row in range(merge_start_row, merge_end_row + 1):
            if sheet.row_dimensions[row].height is not None:
                total_height += sheet.row_dimensions[row].height
            else:
                total_height += sheet.sheet_format.defaultRowHeight

        # Add images into merged column E and resize them to fit within the merged cells
        color_code_cell = sheet[f'A{merge_start_row}'].value
        image_path = find_image(color_code_cell, image_folder)

        if image_path:
            print(f"Adding image {os.path.basename(image_path)} at path {image_path} to row {merge_start_row}")
            img = Image(image_path)
            img.width = 250
            img.height = 100
            img.anchor = f'E{merge_start_row}'
            sheet.add_image(img)
        else:
            print(f"Image starting with {str(color_code_cell)[:4]} not found in {image_folder}")

        i += 1

    # Save the modified workbook as a new file
    wb.save(new_excel_path)
    print(f"Rows added, images added, and workbook saved successfully at {new_excel_path}")

# Define paths
excel_path = 'MARKETING_SIRA_STOCK.xlsx'
new_excel_path = 'SIRA_STOCK_WITH_IMAGES_final.xlsx'
image_folder = '/content/drive/MyDrive/dbms/CATALOGUE'

# Add rows around color codes, merge cells in column E, and add images
add_rows_around_color_codes(excel_path, new_excel_path, image_folder)